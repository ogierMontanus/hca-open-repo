"""hca_v092_to_csv.py — ingest the V0.92 nine-workbook source into the same
three normalized CSVs that hca_xlsx_to_csv.py produces from V0.82.

V0.92 ships Calendar, Persons, Places (Steder) and Diaries; the Works
register is still V0.82-only. This script therefore emits **persons +
places** into entities.csv and the V0.92 diary into diary.csv/references.csv.
Run the V0.82 ingester in parallel until Works land in a future V0.9x.

Output goes to data/normalized_v092/ so the V0.82 outputs at
data/normalized/ stay untouched (the mockup still consumes those).

ID scheme — distinct from V0.82's "Reg…" so the two sets don't collide:
  - persons  : "P{PerID:05d}"           (8,918 dim rows; FactDiaPerPag link)
  - places   : "L{LocID:05d}"           (2,436 dim rows; FactDiaLoc link)
  - diary    : "Pag{Vol:02d}{Page:04d}" matching V0.82 convention
  - page-id  : derived from DiaPagID via DimDiaPag (VolRef, PageRef)

Schema mapping is documented in docs/data-model/v0.92-structural-diff.md.

Usage (PowerShell on Windows):
  python scripts/normalization/hca_v092_to_csv.py
"""

import csv
import os
import re
import sys
import warnings
from pathlib import Path

warnings.filterwarnings("ignore")

try:
    from openpyxl import load_workbook
except ImportError:
    sys.exit("openpyxl is required:  pip install openpyxl")


ROOT = Path(__file__).resolve().parents[2]
SRC_DIR = ROOT / "data" / "raw" / "HCA REPOSITORY V0.92"
OUT_DIR = ROOT / "data" / "normalized_v092"

FILES = {
    "person":   SRC_DIR / "PersonData-PQ-V0.92.xlsx",
    "location": SRC_DIR / "LocationData-PQ-V0.92.xlsx",
    "diary":    SRC_DIR / "DiaryData-PQ-V0.92.xlsx",
    "factdim":  SRC_DIR / "DiaryFactDim-PQ-V0.92.xlsx",
}

ENTITY_FIELDS = [
    "entity_id", "entity_type", "category_h1", "genre_h2",
    "form_h3", "subform_h4", "label", "description",
    "see", "see_also", "year_derived", "date_derived", "person_derived",
]
DIARY_FIELDS = ["vol", "page", "date", "month", "year", "heading", "text"]
REF_FIELDS = ["page_id", "entity_id", "entity_label", "vol", "page", "seq"]


def sheet_rows(path: Path, sheet: str):
    """Yield dicts keyed by header for one sheet."""
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb[sheet]
        it = ws.iter_rows(values_only=True)
        header = list(next(it))
        for row in it:
            if all(c is None for c in row):
                continue
            yield dict(zip(header, row))
    finally:
        wb.close()


def s(v):
    """Stringify a cell, treating None / NaN as empty."""
    if v is None:
        return ""
    return str(v).strip()


def person_id(per_id) -> str:
    """Stable entity_id for a person, from PerID."""
    return f"P{int(per_id):05d}"


def place_id(loc_id) -> str:
    """Stable entity_id for a place, from LocID."""
    return f"L{int(loc_id):05d}"


# Roman volume → integer for the Pag handle. V0.82's Pag convention uses
# decimal volume numbers (Pag010001 = Vol I, page 1). Mirror that here so
# downstream builders that expect Pag-handle padding still parse.
ROMAN_TO_INT = {
    "I": 1, "II": 2, "III": 3, "IV": 4, "V": 5, "VI": 6,
    "VII": 7, "VIII": 8, "IX": 9, "X": 10, "XI": 11, "XII": 12,
    "XIII": 13, "XIV": 14, "XV": 15, "XVI": 16, "XVII": 17,
    "XVIII": 18, "XIX": 19, "XX": 20, "XXI": 21, "XXII": 22,
    "XXIII": 23, "XXIV": 24,
}


def vol_to_int(vol_ref) -> int:
    """Map 'VI' or 6 → 6. Unknown values → 0 so the row is still emittable
    and shows up in audits."""
    if vol_ref is None:
        return 0
    if isinstance(vol_ref, int):
        return vol_ref
    sv = str(vol_ref).strip().upper()
    if sv in ROMAN_TO_INT:
        return ROMAN_TO_INT[sv]
    # Already arabic?
    try:
        return int(sv)
    except ValueError:
        return 0


def page_handle(vol_ref, page_ref) -> str:
    """V0.82 'Pag010001' style handle, used as references.page_id."""
    v = vol_to_int(vol_ref)
    try:
        p = int(page_ref)
    except (TypeError, ValueError):
        p = 0
    return f"Pag{v:02d}{p:04d}"


def normalize_date(year, month_no, day_no):
    """Render YYYY-MM-DD with X stand-ins where the parts are missing.
    Matches the V0.82 'YYYY-XX-XX' convention seen in diary.csv."""
    y = s(year) or "XXXX"
    try:
        mo = int(month_no)
        m = f"{mo:02d}"
    except (TypeError, ValueError):
        m = "XX"
    try:
        d = int(day_no)
        dd = f"{d:02d}"
    except (TypeError, ValueError):
        dd = "XX"
    return f"{y}-{m}-{dd}"


# ----------------------------------------------------------------------------

def build_entities():
    """Persons + places → entities.csv shape.

    Persons: DimPer1 (in DiaryFactDim) is the canonical, FK-keyed subset of
    the register that's referenced from Vol VI/VII pages — 8,918 rows. Use
    it instead of PersonData.RawPer1 (10,229) so PerID matches the fact
    tables. Persons outside DimPer1 don't yet have structured references in
    V0.92 anyway.

    Same logic for places via DimLoc1 (2,436 rows)."""
    out = []

    for r in sheet_rows(FILES["factdim"], "DimPer1"):
        per_id = r.get("PerID")
        if per_id is None:
            continue
        yob, yod = s(r.get("YearOfBirth")), s(r.get("YearOfDeath"))
        # Render lifespan into year_derived like V0.82 did for works
        if yob and yod:
            year_derived = f"{yob}–{yod}"
        else:
            year_derived = yob or yod
        out.append({
            "entity_id":    person_id(per_id),
            "entity_type":  "person",
            "category_h1":  "PERSON-REGISTER",
            "genre_h2":     "",
            "form_h3":      "",
            "subform_h4":   "",
            "label":        s(r.get("RegistryTitle")),
            "description":  s(r.get("RegistryDescription")),
            "see":          "",   # to be filled from a future Raw.See-Also
            "see_also":     "",
            "year_derived": year_derived,
            "date_derived": "",
            "person_derived": "",
        })

    for r in sheet_rows(FILES["factdim"], "DimLoc1"):
        loc_id = r.get("LocID")
        if loc_id is None:
            continue
        country = s(r.get("Country"))
        region = s(r.get("Region"))
        out.append({
            "entity_id":    place_id(loc_id),
            "entity_type":  "place",
            "category_h1":  "STED-REGISTER",
            "genre_h2":     "",
            "form_h3":      "",
            "subform_h4":   "",
            "label":        s(r.get("LocationTitle")),
            "description":  " · ".join([x for x in (country, region) if x]),
            "see":          "",
            "see_also":     "",
            "year_derived": "",
            "date_derived": "",
            # Repurpose person_derived to carry lat,lon when present, so the
            # column isn't wasted on places. Documented in the diff doc.
            "person_derived": (
                f"{s(r.get('Lat'))},{s(r.get('Lon'))}" if r.get("Lat") else ""
            ),
        })

    return out


def build_diary():
    """DimDiaPag2 → diary.csv shape. One row per diary page."""
    out = []
    for r in sheet_rows(FILES["factdim"], "DimDiaPag2"):
        if not r.get("DiaPagID"):
            continue
        v_int = vol_to_int(r.get("VolRef"))
        # Existing diary.csv uses 'I', 'VII' Roman numerals in vol — preserve
        roman = s(r.get("VolRef"))
        # Date pieces
        # V0.92 has a single 'Date' cell (datetime), Month and Year cols
        d = r.get("Date")
        if hasattr(d, "year"):
            date_str = d.strftime("%Y-%m-%d")
            month_str = f"{d.month:02d}"
            year_str = str(d.year)
        else:
            month_str = s(r.get("Month"))
            year_str = s(r.get("Year"))
            date_str = normalize_date(year_str, month_str, None)
        out.append({
            "vol":     roman,
            "page":    s(r.get("PageRef")),
            "date":    date_str,
            "month":   month_str,
            "year":    year_str,
            "heading": s(r.get("DiaryDayHeading")),
            "text":    s(r.get("DiaryTextLines")),
        })
    return out


def build_references():
    """FactDiaPerPag + FactDiaLocPag → references.csv shape.

    Join paths:
      page  = DimDiaPag.DiaPagID → (VolRef, PageRef) → page_handle()
      person= FactDiaPerPag.PerID + DimPer1.RegistryTitle
      place = FactDiaLocPag.LocID + DimLoc1.LocationTitle
    """
    # Page index: DiaPagID → (vol_roman, page_ref)
    page_index = {}
    for r in sheet_rows(FILES["factdim"], "DimDiaPag2"):
        page_index[r["DiaPagID"]] = (s(r.get("VolRef")), s(r.get("PageRef")))

    # Person labels
    per_labels = {}
    for r in sheet_rows(FILES["factdim"], "DimPer1"):
        per_labels[r["PerID"]] = s(r.get("RegistryTitle"))

    # Place labels
    loc_labels = {}
    for r in sheet_rows(FILES["factdim"], "DimLoc1"):
        loc_labels[r["LocID"]] = s(r.get("LocationTitle"))

    out = []
    seq_counter = {}   # page_id → running per-page seq

    def emit(page_id, vol_roman, page_ref, ent_id, ent_label):
        seq_counter[page_id] = seq_counter.get(page_id, 0) + 1
        out.append({
            "page_id":      page_id,
            "entity_id":    ent_id,
            "entity_label": ent_label,
            "vol":          vol_roman,
            "page":         page_ref,
            "seq":          seq_counter[page_id],
        })

    for r in sheet_rows(FILES["factdim"], "FactDiaPerPag"):
        dia_pag_id = r.get("DiaPagID")
        per_id = r.get("PerID")
        if dia_pag_id not in page_index or per_id is None:
            continue
        vol_roman, page_ref = page_index[dia_pag_id]
        emit(page_handle(vol_roman, page_ref), vol_roman, page_ref,
             person_id(per_id), per_labels.get(per_id, ""))

    for r in sheet_rows(FILES["factdim"], "FactDiaLocPag"):
        dia_pag_id = r.get("DiaPagID")
        loc_id = r.get("LocID")
        if dia_pag_id not in page_index or loc_id is None:
            continue
        vol_roman, page_ref = page_index[dia_pag_id]
        emit(page_handle(vol_roman, page_ref), vol_roman, page_ref,
             place_id(loc_id), loc_labels.get(loc_id, ""))

    return out


def write_csv(path, fields, rows):
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as f:
        w = csv.DictWriter(f, fieldnames=fields)
        w.writeheader()
        w.writerows(rows)


def main():
    for kind, p in FILES.items():
        if not p.exists():
            sys.exit(f"V0.92 source missing: {p}")

    print(f"Ingesting V0.92 from {SRC_DIR.relative_to(ROOT)}")

    ents = build_entities()
    persons = sum(1 for r in ents if r["entity_type"] == "person")
    places = sum(1 for r in ents if r["entity_type"] == "place")
    write_csv(OUT_DIR / "entities.csv", ENTITY_FIELDS, ents)
    print(f"  entities.csv: {len(ents):>6} rows  ({persons} persons + {places} places)")

    diary = build_diary()
    write_csv(OUT_DIR / "diary.csv", DIARY_FIELDS, diary)
    print(f"  diary.csv:    {len(diary):>6} rows")

    refs = build_references()
    write_csv(OUT_DIR / "references.csv", REF_FIELDS, refs)
    print(f"  references.csv: {len(refs):>6} rows")

    print(f"Done. Output -> {OUT_DIR.relative_to(ROOT)}")


if __name__ == "__main__":
    main()
