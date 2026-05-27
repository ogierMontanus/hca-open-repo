#!/usr/bin/env python3
"""
hca_xlsx_to_csv.py
------------------
Reads HCA-Repository V0.82.xlsx (or equivalent CSV exports) and writes
normalized CSVs to data/normalized/:

  entities.csv    — persons, places, works from Registry sheet
  diary.csv       — diary entries from Diary sheet
  references.csv  — diary-page → entity joins from RefInDiaryPage sheet

Usage:
  python3 scripts/normalization/hca_xlsx_to_csv.py
  python3 scripts/normalization/hca_xlsx_to_csv.py --input path/to/file.xlsx
  python3 scripts/normalization/hca_xlsx_to_csv.py --input path/to/registry.csv

Supports both .xlsx and .csv input (base formats).
"""

import argparse
import csv
import os
import sys

XLSX_DEFAULT = os.path.join(
    os.path.dirname(__file__), "..", "..", "data", "raw", "HCA-Repository V0.82.xlsx"
)
OUT_DIR = os.path.join(os.path.dirname(__file__), "..", "..", "data", "normalized")

# ── Sheet / column mappings ──────────────────────────────────────────────────

# Registry sheet columns (row 1 = header):
#   PKRegistryTitelID, RegistryCategory (H1), WorRegSubCat.WorkGenre (H2),
#   WorRegSubCat.RegistryForm (H3), WorRegSubCat.WorkSubForm (H4),
#   RegistryTitle, RegistryDescription, SeeTitle, SeeAlsoTittle,
#   YearDerived, DateDerived, PersonDerived

# Diary sheet columns (row 1 = header):
#   VolRef, Date, Month, Year, PageRef, DiaryDayHeading, DiaryTextLines

# RefInDiaryPage sheet columns (row 1 = header):
#   PKRegistryPageID, FKRegistryTitelID, RegistryTitel, VolRef, PageRef, BookSeqNo

CATEGORY_MAP = {
    "PERSON-REGISTER": "person",
    "STED-REGISTER":   "place",
    "VÆRK-REGISTER":   "work",
}

# ── Helpers ──────────────────────────────────────────────────────────────────

def load_xlsx(path):
    try:
        import openpyxl
    except ImportError:
        sys.exit("openpyxl not installed. Run: pip install openpyxl")
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheets = {}
    for name in wb.sheetnames:
        ws = wb[name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        header = [str(c).strip() if c else "" for c in rows[0]]
        sheets[name] = [dict(zip(header, row)) for row in rows[1:] if any(row)]
    wb.close()
    return sheets


def load_csv_dir(path):
    """Accept a directory of CSVs named after sheets, or a single CSV file."""
    sheets = {}
    if os.path.isdir(path):
        for fn in os.listdir(path):
            if fn.endswith(".csv"):
                name = fn[:-4]
                with open(os.path.join(path, fn), encoding="utf-8-sig") as f:
                    sheets[name] = list(csv.DictReader(f))
    elif os.path.isfile(path) and path.endswith(".csv"):
        name = os.path.splitext(os.path.basename(path))[0]
        with open(path, encoding="utf-8-sig") as f:
            sheets[name] = list(csv.DictReader(f))
    return sheets


def write_csv(path, fieldnames, rows):
    os.makedirs(os.path.dirname(path), exist_ok=True)
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=fieldnames, extrasaction="ignore")
        w.writeheader()
        w.writerows(rows)
    print(f"  Wrote {len(rows):,} rows → {os.path.relpath(path)}")

# ── Normalizers ──────────────────────────────────────────────────────────────

def normalize_registry(rows):
    """Registry → entities.csv"""
    out = []
    for r in rows:
        pk = r.get("PKRegistryTitelID") or r.get("PKRegistryTitelID ")
        if not pk:
            continue
        cat_raw = r.get("RegistryCategory (H1)", "") or ""
        entity_type = CATEGORY_MAP.get(cat_raw.strip(), "unknown")
        out.append({
            "entity_id":    pk,
            "entity_type":  entity_type,
            "category_h1":  cat_raw.strip(),
            "genre_h2":     (r.get("WorRegSubCat.WorkGenre (H2)") or "").strip(),
            "form_h3":      (r.get("WorRegSubCat.RegistryForm (H3)") or "").strip(),
            "subform_h4":   (r.get("WorRegSubCat.WorkSubForm (H4)") or "").strip(),
            "label":        (r.get("RegistryTitle") or "").strip(),
            "description":  (r.get("RegistryDescription") or "").strip(),
            "see":          (r.get("SeeTitle") or "").strip(),
            "see_also":     (r.get("SeeAlsoTittle") or "").strip(),
            "year_derived": (r.get("YearDerived") or "").strip(),
            "date_derived": (r.get("DateDerived") or "").strip(),
            "person_derived": (r.get("PersonDerived") or "").strip(),
        })
    return out


def normalize_diary(rows):
    """Diary sheet → diary.csv"""
    out = []
    seen = set()
    for r in rows:
        vol  = r.get("VolRef", "")
        page = r.get("PageRef", "")
        key  = (vol, page)
        if not vol and not page:
            continue
        if key not in seen:
            seen.add(key)
        out.append({
            "vol":       str(vol).strip() if vol else "",
            "page":      str(page).strip() if page else "",
            "date":      (r.get("Date") or "").strip(),
            "month":     (r.get("Month") or "").strip(),
            "year":      str(r.get("Year") or "").strip(),
            "heading":   (r.get("DiaryDayHeading") or "").strip(),
            "text":      (r.get("DiaryTextLines") or "").strip(),
        })
    return out


def normalize_refs(rows):
    """RefInDiaryPage → references.csv"""
    out = []
    for r in rows:
        page_id = r.get("PKRegistryPageID")
        reg_id  = r.get("FKRegistryTitelID")
        if not page_id or not reg_id:
            continue
        out.append({
            "page_id":      str(page_id).strip(),
            "entity_id":    str(reg_id).strip(),
            "entity_label": (r.get("RegistryTitel") or "").strip(),
            "vol":          str(r.get("VolRef") or "").strip(),
            "page":         str(r.get("PageRef") or "").strip(),
            "seq":          str(r.get("BookSeqNo") or "").strip(),
        })
    return out

# ── Main ─────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Normalize HCA xlsx/csv to entities, diary, and references CSVs.")
    parser.add_argument("--input", default=XLSX_DEFAULT, help="Path to .xlsx file or CSV directory")
    parser.add_argument("--out", default=OUT_DIR, help="Output directory for normalized CSVs")
    args = parser.parse_args()

    inp = os.path.normpath(args.input)
    if not os.path.exists(inp):
        sys.exit(f"Input not found: {inp}")

    print(f"Reading: {inp}")
    if inp.endswith(".xlsx"):
        sheets = load_xlsx(inp)
    else:
        sheets = load_csv_dir(inp)

    os.makedirs(args.out, exist_ok=True)

    # entities.csv
    if "Registry" in sheets:
        rows = normalize_registry(sheets["Registry"])
        write_csv(
            os.path.join(args.out, "entities.csv"),
            ["entity_id", "entity_type", "category_h1", "genre_h2", "form_h3",
             "subform_h4", "label", "description", "see", "see_also",
             "year_derived", "date_derived", "person_derived"],
            rows,
        )
    else:
        print("  Warning: 'Registry' sheet not found — skipping entities.csv")

    # diary.csv
    if "Diary" in sheets:
        rows = normalize_diary(sheets["Diary"])
        write_csv(
            os.path.join(args.out, "diary.csv"),
            ["vol", "page", "date", "month", "year", "heading", "text"],
            rows,
        )
    else:
        print("  Warning: 'Diary' sheet not found — skipping diary.csv")

    # references.csv
    if "RefInDiaryPage" in sheets:
        rows = normalize_refs(sheets["RefInDiaryPage"])
        write_csv(
            os.path.join(args.out, "references.csv"),
            ["page_id", "entity_id", "entity_label", "vol", "page", "seq"],
            rows,
        )
    else:
        print("  Warning: 'RefInDiaryPage' sheet not found — skipping references.csv")

    print("Done.")


if __name__ == "__main__":
    main()
