#!/usr/bin/env python3
"""Build Excel workbook with the 80 verified copilot cases + full SV14 register
classification, per the markup instruction in place-typology.md section G."""

import xml.etree.ElementTree as ET
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

NS = {'tei': 'http://www.tei-c.org/ns/1.0'}

CATEGORY_NAMES = {
    'bebygget': 'Bebyggede områder',
    'admreg': 'Lande, administrative enheder og regioner',
    'vand': 'Vandområder',
    'landskab': 'Landskabsformer og naturfænomener',
    'anlaeg': 'Bygninger, anlæg og fortidsminder',
    'park': 'Parker, haver og naturområder',
    'usikker': 'USIKKER',
}
CATEGORY_NUM = {
    'bebygget': 1, 'admreg': 2, 'vand': 3, 'landskab': 4, 'anlaeg': 5, 'park': 6, 'usikker': '',
}

# GeoNames class -> default slug
CLASS_DEFAULT = {
    'P': 'bebygget',
    'A': 'admreg',
    'H': 'vand',
    'L': 'landskab',
    'T': 'landskab',
    'S': 'anlaeg',
    'R': 'anlaeg',
    'U': 'vand',
    'V': 'landskab',
}

# Explicit per-code overrides (regel 2 i afsnit G)
CODE_OVERRIDE = {
    'A.ADM3': 'bebygget', 'A.ADM3H': 'bebygget',
    'A.ADM4': 'bebygget', 'A.ADM4H': 'bebygget',
    'A.ADM5': 'bebygget', 'A.ADM5H': 'bebygget',
    'L.PRK': 'park',
    'L.RESN': 'park', 'L.RES': 'park', 'L.RESA': 'park', 'L.RESF': 'park',
    'L.RESH': 'park', 'L.RESP': 'park', 'L.RESV': 'park', 'L.RESW': 'park',
    'L.RGN': 'admreg', 'L.RGNH': 'admreg', 'L.RGNE': 'admreg', 'L.RGNL': 'admreg',
    'S.CAVE': 'landskab',
    'S.GDN': 'park',
}

# Navngivne enkelttilfælde (regel 1, højeste prioritet) -- keyed by xml:id
NAMED_CASES = {
    'geo-982821': ('usikker', 'Lilienstein (fejlkoblet post)',
        'S.FRM peger paa farm/Mpumalanga, Sydafrika; det tilsigtede sted er '
        'formentlig bordbjerget i Sachsisk Schweiz -> landskab efter rettelse. '
        'Se place-typology.md afsnit C punkt 5 og afsnit G.'),
}
NAMED_CASES_BY_NAME = {
    'Hippodrome of Constantinople': ('anlaeg',
        'S.GDN (garden) er en GeoNames-fejlkodning; hippodromet er et antikt '
        'monument/stadion, ikke en have. Regel 1 tilsidesaetter regel 2s '
        'have-undtagelse.'),
}

# Regel 4: de 90 PPL-placeholder-poster er individuelt gennemgået (ikke
# navnemønster-regex), fordi stikprøvekontrol viste, at et generisk
# "default = bebygget" var direkte forkert her: PPL-koden i dette register
# bruges næsten udelukkende til kirker, slotte, teatre, monumenter og
# landskabsformer, som en GeoNames-matchning ikke er lykkedes for -- IKKE
# til bebyggelser (jf. afsnit A's egen kvalitative beskrivelse). Se
# selvkritik-noten i README/rapport-arket.
PPL_NAME_OVERRIDE = {
    "Odense Adelige Jomfrukloster": ("anlaeg", "Kloster."),
    "Klokkedybet": ("vand", "Dyb/vandhul i Odense Å."),
    "Nonnebakken": ("anlaeg", "Vikingeringborg, Odense — fortidsminde (jf. afsnit A)."),
    "Munke-Mose": ("park", "Anlagt bypark i Odense, trods navnets \"mose\"-led."),
    "Hunderup Skov": ("park", "Navngivet park-eksempel i afsnit B."),
    "Skt. Knuds Kirke": ("anlaeg", "Kirke."),
    "Dalum": ("bebygget", "Bydel/lokalitet ved Odense."),
    "Sankt Knuds Kirke": ("anlaeg", "Kirke (variant af Skt. Knuds Kirke)."),
    "Vor Frue Kirke": ("anlaeg", "Kirke."),
    "Næsbyhoved Skov": ("park", "Navngivet park-eksempel i afsnit B."),
    "Lübeck Rådhus": ("anlaeg", "Rådhus."),
    "den botaniske Have": ("park", "Navngivet park-eksempel i afsnit B."),
    "Falleberthor": ("usikker", "Uidentificerbart sted, jf. afsnit C punkt 1."),
    "Der Braunschweiger Dom": ("anlaeg", "Domkirke."),
    "Burg Dankwarderode": ("anlaeg", "Borg."),
    "Zwinger Tower": ("anlaeg", "Slotsanlæg (Dresden Zwinger)."),
    "Hovedkirken ": ("anlaeg", "Kirke."),
    "Kloster Ilsenburg": ("anlaeg", "Kloster."),
    "Ilse dalen": ("landskab", "Dal (Ilsetal, Harzen)."),
    "Bloksbjerg": ("landskab", "Bjerg (Brocken, Harzen)."),
    "Teufelsmauer": ("landskab", "Naturlig klippeformation."),
    "Baumannshöhle": ("landskab", "Grotte (naturlig)."),
    "Schloss Blankenburg": ("anlaeg", "Slot."),
    "Gellerts Grav": ("anlaeg", "Gravsted (digteren Gellert) — fortidsminde."),
    "Hôtel de Bavière": ("anlaeg", "Hotel."),
    "Leipziger Altes Theater": ("anlaeg", "Teater."),
    "Dom zu Meißen": ("anlaeg", "Domkirke."),
    "Albrechtsburg Meissen": ("anlaeg", "Slot/borg."),
    "Augustusbrücke": ("anlaeg", "Bro."),
    "det kongelige Theater": ("anlaeg", "Teater. Identifikations-usikkerhed, jf. afsnit C punkt 9."),
    "Linchkeschen Bade": ("anlaeg", "Bad/spa-etablissement. Identifikations-usikkerhed, jf. afsnit C punkt 9."),
    "Das grüne Gewölb": ("anlaeg", "Museum (Grünes Gewölbe, Dresden)."),
    "Schloß Lohmen": ("anlaeg", "Slot."),
    "Dorfkirche Lohmen": ("anlaeg", "Landsbykirke."),
    "Teufelsküche": ("landskab", "Naturlig klippeformation."),
    "Felsenburg Neurathen": ("anlaeg", "Klippeborg/ruin (Sachsisk Schweiz)."),
    "Teufelsbrücke": ("anlaeg", "Bro (menneskeskabt, trods sagnnavnet)."),
    "St-Johannis-Kirche": ("anlaeg", "Kirke."),
    "Sneiderloch": ("landskab", "Grotte/klippehule (Harzen)."),
    "Böhmen": ("admreg", "Historisk region (Bøhmen)."),
    "Stadtkirche St. Marien": ("anlaeg", "Kirke."),
    "Plauenscher Grund": ("landskab", "Dal/kløft ved Dresden."),
    "Der Kaiser von Rusland": ("anlaeg", "Sandsynligvis værtshus/hotel. Identifikations-usikkerhed, jf. afsnit C punkt 9."),
    "Operahuset Webers": ("anlaeg", "Operahus. Identifikations-usikkerhed, jf. afsnit C punkt 9."),
    "Königstätisches Theater": ("anlaeg", "Teater. Identifikations-usikkerhed, jf. afsnit C punkt 9."),
    "Konzerthaus Berlin": ("anlaeg", "Koncerthus."),
    "Museet ": ("anlaeg", "Museum. Identifikations-usikkerhed, jf. afsnit C punkt 9."),
    "Schloss Ludwigslust": ("anlaeg", "Slot."),
    "Italien ": ("admreg", "Land."),
    "Accademia di San Luca": ("anlaeg", "Kunstakademi, Rom."),
    "Tomba di Virgilio": ("anlaeg", "Gravsted (Vergil) — fortidsminde."),
    "Klosteret St. Antonio": ("anlaeg", "Kloster. Identifikations-usikkerhed, jf. afsnit C punkt 9."),
    "Carlo Felice": ("anlaeg", "Operahus, Genova (Teatro Carlo Felice)."),
    "Teatro Pallacorda": ("anlaeg", "Teater. Identifikations-usikkerhed, jf. afsnit C punkt 9."),
    "Teatro Alibert": ("anlaeg", "Teater, Rom."),
    "Teatro di Apollo": ("anlaeg", "Teater, Rom."),
    "Teatro dei Fiorentini": ("anlaeg", "Teater, Napoli."),
    "Teatro Fenise": ("anlaeg", "Teater. Identifikations-usikkerhed, jf. afsnit C punkt 9."),
    "Kongens Nytorv": ("anlaeg", "Navngivet plads-eksempel i afsnit B."),
    "Schloss Breitenburg": ("anlaeg", "Slot (jf. Breitenburg i den 80-verificerede sample)."),
    "Schöner Brunnen": ("anlaeg", "Springvand/monument, Nürnberg."),
    "Sebalduskirche": ("anlaeg", "Kirke, Nürnberg."),
    "Hofkirche Innsbruck": ("anlaeg", "Kirke."),
    "Palazzo degli Uffizi": ("anlaeg", "Palads/museum (Uffizierne), Firenze."),
    "Valico di Somma": ("landskab", "Bjergpas."),
    "monte mario": ("landskab", "Bakke/høj, Rom."),
    "Palazzo Borghese": ("anlaeg", "Palads, Rom."),
    "Via di Ripetta": ("anlaeg", "Gade, Rom."),
    "Chiesa di Sant'Antonio Abate all'Esquilino": ("anlaeg", "Kirke, Rom."),
    "Grotterne": ("landskab", "Grotter (naturlige)."),
    "Grande Cascata di Tivoli": ("landskab", "Vandfald, Tivoli."),
    "Via della Purificazione": ("anlaeg", "Gade, Rom."),
    "Santa Maria in Traspontina": ("anlaeg", "Kirke, Rom."),
    "Mola di Gaeta": ("bebygget", "Lokalitet/by ved Gaeta (nu del af Formia)."),
    "Teatro di San Carlo": ("anlaeg", "Operahus, Napoli."),
    "Piazza Municipio": ("anlaeg", "Plads, Napoli."),
    "Chiesa di Santa Maria della Mercede a Montecalvario": ("anlaeg", "Kirke, Napoli."),
    "Katidral ta’ San Pawl": ("anlaeg", "Katedral, Malta."),
    "Ermou": ("anlaeg", "Gade, Athen (Ermou-gaden)."),
    "Themistocles Tomb": ("anlaeg", "Gravsted — fortidsminde."),
    "Galata Mevlevihanesi Müzesi": ("anlaeg", "Tidligere dervish-kloster, nu museum (jf. grænsetilfælde i afsnit B)."),
    "Mısır Çarşısı": ("anlaeg", "Basar/markedsbygning, Istanbul."),
    "Emirgan Mosque": ("anlaeg", "Moske."),
    "Kız Kulesi": ("anlaeg", "Tårn/fortidsminde, jf. afsnit C punkt 4."),
    "Ovidius Kulesi": ("anlaeg", "Tårn (Ovid-associeret), formentlig Constanța."),
    "Canalul Dunăre-Marea Neagră": ("vand", "Kanal (Donau-Sortehavskanalen) — kanalundtagelsen, afsnit F."),
    "Trajan's Plaque": ("anlaeg", "Mindetavle/monument."),
    "Golubac Fortress": ("anlaeg", "Fæstning."),
    "Petrovaradin fortress": ("anlaeg", "Fæstning."),
}


def name_pattern_fallback(name: str):
    """Sidste udvej for @type-værdier, der hverken er en kendt GeoNames-kode
    eller en af de 90 opslåede PPL-navne (fx fremtidige høstede poster)."""
    n = name.lower()
    if 'kanal' in n or 'canal' in n:
        return 'vand', 'Navnemønster: "kanal/canal" — kanalundtagelsen (afsnit F)'
    if n.startswith('via ') or ' street' in n or 'straße' in n or 'strasse' in n:
        return 'anlaeg', 'Navnemønster: gade/straße'
    return 'usikker', 'Intet navnemønster matchede — kræver manuel gennemgang'


def classify(xml_id, name, geotype, note_text):
    # Regel 1: navngivne enkelttilfælde
    if xml_id in NAMED_CASES:
        slug, _, reason = NAMED_CASES[xml_id]
        return slug, 'Navngivet enkelttilfælde', reason
    if name in NAMED_CASES_BY_NAME:
        slug, reason = NAMED_CASES_BY_NAME[name]
        return slug, 'Navngivet enkelttilfælde', reason

    # Regel 2/3: kodetabel
    if geotype and geotype not in (None, 'null', 'PPL'):
        if geotype in CODE_OVERRIDE:
            return CODE_OVERRIDE[geotype], 'Kodeundtagelse (regel 2)', f'{geotype} -> eksplicit undtagelse'
        cls = geotype.split('.')[0]
        if cls in CLASS_DEFAULT:
            return CLASS_DEFAULT[cls], 'GeoNames-klasse (regel 3)', f'{geotype} -> klasse {cls} default'
        return 'usikker', 'Ukendt GeoNames-klasse', f'{geotype} har uklassificeret klassepræfiks'

    # Regel 4: navnemønster / manglende type
    if geotype == 'PPL':
        if name in PPL_NAME_OVERRIDE:
            slug, reason = PPL_NAME_OVERRIDE[name]
            return slug, 'PPL-placeholder, individuelt gennemgået', reason
        slug, reason = name_pattern_fallback(name)
        return slug, 'PPL-placeholder, navnemønster (uden for opslagstabel)', reason
    if geotype == 'null' or geotype is None:
        if note_text and 'village' in note_text.lower():
            return 'bebygget', 'Manuel (fra notetekst, regel 4)', 'Notetekst beskriver stedet som "small coastal village"'
        slug, reason = name_pattern_fallback(name)
        return slug, 'Manglende @type (regel 4)', reason

    return 'usikker', 'Uafklaret (regel 5)', 'Intet af reglerne 1-4 gav et forsvarligt resultat'


def load_sv14(path):
    tree = ET.parse(path)
    root = tree.getroot()
    places = root.findall('.//tei:place', NS)
    rows = []
    for p in places:
        xml_id = p.get('{http://www.w3.org/XML/1998/namespace}id', '')
        geotype = p.get('type')
        main_name_el = p.find('tei:placeName[@type="main"]', NS)
        name = main_name_el.text if main_name_el is not None and main_name_el.text else '(uden navn)'
        country_el = p.find('tei:country', NS)
        region_el = p.find('tei:region', NS)
        note_el = p.find('tei:note', NS)
        country = country_el.text if country_el is not None and country_el.text else ''
        region = region_el.text if region_el is not None and region_el.text else ''
        note_text = note_el.text if note_el is not None and note_el.text else ''

        slug, method, reason = classify(xml_id, name, geotype, note_text)
        rows.append({
            'xml_id': xml_id,
            'Stednavn': name,
            'Land': country,
            'Region': region,
            'GeoNames @type': geotype if geotype else '(mangler)',
            'Kategori nr.': CATEGORY_NUM.get(slug, ''),
            'Kategori (slug)': slug,
            'Kategori (navn)': CATEGORY_NAMES.get(slug, slug),
            'Metode': method,
            'Begrundelse': reason,
        })
    return rows


# --- De 80 verificerede cases (fra place-categorization-copilot-report.md) ---
VERIFIED_80 = [
    ("Sieghartskirchen", "Kommune/by i Niederösterreich", "bebygget", "Bebyggelsesnavn.", ""),
    ("Gotha", "By i Thüringen", "bebygget", "By.", ""),
    ("Woigwitz", "Historisk landsby i Schlesien; Meyers angiver “Dorf”", "bebygget", "Landsby.", "[17]"),
    ("Malines", "Mechelen/Malines, Belgien", "bebygget", "By.", ""),
    ("Heiligeostedten", "Heiligenstedten, Holsten", "bebygget", "Landsby/kommuneform.", ""),
    ("Randers", "Dansk købstad/by", "bebygget", "By.", ""),
    ("Källtorp", "Svensk lokalitet/gårdnavn", "bebygget", "Bebyggelses-/gårdnavn.", ""),
    ("Trollhättan", "Svensk by", "bebygget", "Bare navneform, ikke “faldene”.", ""),
    ("Karrebæksminde", "Dansk kystby", "bebygget", "Bebyggelse.", ""),
    ("Strengberg", "Østrigsk kommune", "bebygget", "Kommune/by.", ""),
    ("Nordtyskland", "Historisk/geografisk region", "admreg", "Regional betegnelse.", ""),
    ("Reuss", "Reuss-floden i Schweiz", "vand", "Schweiz' turistportal beskriver Reuss som central-schweizisk flod.", "[16]"),
    ("Tostedt", "Tysk kommune/by", "bebygget", "Bebyggelse.", ""),
    ("Horsens", "Dansk by", "bebygget", "By.", ""),
    ("Marmara-Øen", "Ø i Marmarahavet", "landskab", "Ø.", ""),
    ("Valencia", "Valencia, Spanien", "bebygget", "Primært byreferent.", ""),
    ("Mori", "Mori i Trentino", "bebygget", "By/kommune.", ""),
    ("Theresienstadt", "Terezín/Theresienstadt, fæstningsby", "bebygget", "Kilden kalder stedet en lille by og fæstningsby; her valgt som bynavn.", "[5]"),
    ("Butzbach", "Tysk by", "bebygget", "By.", ""),
    ("Hesselagerstenen", "Damestenen/Hesselagerstenen, vandreblok", "landskab", "Danmarks største kendte vandreblok på land.", "[15]"),
    ("Sevilla", "Sevilla, Spanien", "bebygget", "By.", ""),
    ("Falkensten", "Sandsynlig Falkenstein/Falkensten, slot/ruin eller klippested", "anlaeg", "USIKKER: intern registerform findes, ekstern referent ikke sikkert verificeret.", ""),
    ("Trier", "Trier, Tyskland", "bebygget", "By.", ""),
    ("Hofmansgave", "Herregården/godset Hofmansgave", "anlaeg", "VisitNordfyn beskriver Hofmansgave som herregård/gods.", "[14]"),
    ("Rava", "Rava, kroatisk ø i Zadar-arkipelaget", "landskab", "Visit Zadar beskriver Rava som ø.", "[13]"),
    ("Bornholm", "Ø", "landskab", "Ø.", ""),
    ("Niesen", "Bjerg i Schweiz", "landskab", "Bjerg.", ""),
    ("Rheine", "Tysk by", "bebygget", "By.", ""),
    ("Senlis", "Fransk by", "bebygget", "By.", ""),
    ("Dublin", "Irsk hovedstad/by", "bebygget", "By.", ""),
    ("England", "Land/historisk region", "admreg", "Land/region.", ""),
    ("Monte Rosa", "Alpebjerg/massiv", "landskab", "Bjergmassiv.", ""),
    ("Pietra Mala", "Pietramala/Pietra Mala, italiensk lokalitet", "bebygget", "Vej-/lokalitetsnavn.", ""),
    ("Mornaux", "Les Mornaux/Mornaux, belgisk/fransk lokalitet", "bebygget", "Lokalitet; normalform bør kontrolleres.", ""),
    ("Neuhaus (Schweiz)", "Schweizisk lokalitet, sandsynlig Neuhaus", "bebygget", "Bebyggelses-/lokalitetsnavn.", ""),
    ("Hellebæk", "Dansk by/lokalitet", "bebygget", "Bebyggelse.", ""),
    ("Monte Cavo", "Vulkanbjerg i Albanerbjergene", "landskab", "Bjerg/vulkan.", ""),
    ("Roosendaal", "Nederlandsk by", "bebygget", "By.", ""),
    ("Reichenbach", "Sandsynlig by/lokalitet, ikke Reichenbachfaldene", "bebygget", "Bare navneform; falde ville normalt markeres særskilt.", ""),
    ("Meran", "Merano/Meran", "bebygget", "By.", ""),
    ("Wallersee", "Sø i Østrig", "vand", "Sø.", ""),
    ("Hollabrunn (Ober-Hollabrunn)", "Østrigsk by", "bebygget", "By.", ""),
    ("Rabenstein ved Maxen", "Klippe-/naturpunkt ved Maxen", "landskab", "“ved Maxen” peger mod lokalt topografisk punkt; dog vanskelig.", ""),
    ("Indien", "Land/region", "admreg", "Land.", ""),
    ("Hermupolis Syra", "Ermoupoli på Syros", "bebygget", "By.", ""),
    ("Bellevue (Klampenborg)", "Bellevue Strand/strandbad ved Klampenborg", "park", "700 m sandstrand i Klampenborg.", "[12]"),
    ("Haag", "Den Haag", "bebygget", "By.", ""),
    ("Arbesau", "Historisk bøhmisk/tysk lokalitet", "bebygget", "Registerform; by-/lokalitetsnavn.", ""),
    ("Kowicz", "Sandsynlig polsk lokalitet/by", "bebygget", "Registerform; normalform usikker, kategori stabil.", ""),
    ("Sæby Løve H Holbæk A", "Sæby i Løve Herred, Holbæk Amt", "bebygget", "Lokalitet/sogn angivet med herred/amt.", ""),
    ("Minneapolis", "By i USA", "bebygget", "By.", ""),
    ("Weissenburg Bayern", "Weißenburg i Bayern", "bebygget", "By.", ""),
    ("Rieti", "Italiensk by", "bebygget", "By.", ""),
    ("Danmark Sogn i Uppsala Len", "Danmark sogn, Uppsala län", "admreg", "Sogn = administrativ/kirkelig enhed.", "[6]"),
    ("Mornex", "Fransk kommune/lokalitet", "bebygget", "Bebyggelse.", ""),
    ("Irland", "Land", "admreg", "Land.", ""),
    ("Jerichow", "Tysk by/lokalitet", "bebygget", "By.", ""),
    ("Seinen", "Seine-floden", "vand", "Flod.", ""),
    ("Oldenzaal", "Nederlandsk by", "bebygget", "By.", ""),
    ("Elmshorn", "Tysk by", "bebygget", "By.", ""),
    ("Helvetes-Faldene Trollhätta", "Vandfald ved Trollhättan", "landskab", "Navnet markerer fald/naturfænomen.", ""),
    ("Orange", "Orange, Provence", "bebygget", "By.", ""),
    ("Steinsdorf", "Tysk/østrigsk landsby-/lokalitetsnavn", "bebygget", "Bebyggelse.", ""),
    ("Erfurt", "Tysk by", "bebygget", "By.", ""),
    ("Kuhstall", "Naturlig klippeport i Sachsisk Schweiz", "landskab", "Naturlig sandstens-klippeport.", "[11]"),
    ("Petershøi", "Henriques' Petershøi ved Klampenborg", "anlaeg", "Behandles som ejendom/villa.", "[109]"),
    ("Mägdesprung", "By/lokalitet i Harzen", "bebygget", "Bebyggelsesnavn.", ""),
    ("Münchberg", "Tysk by", "bebygget", "By.", ""),
    ("Semendria (Smederevo)", "Smederevo/Semendria", "bebygget", "By; fæstning sekundær her.", ""),
    ("Galli Li", "Li Galli/Sirenuse-øerne", "landskab", "Tre små øer/islets.", "[8]"),
    ("Ekeberg Christiania", "Ekeberg ved Christiania/Oslo", "landskab", "Høj/terræn- og udsigtområde.", ""),
    ("Mose1", "Sandsynlig Mosel; OCR/normaliseringsproblem", "vand", "“Mose1” læses sandsynligvis Mosel.", "[4]"),
    ("Hessingen (Essingen)", "Essingen/Hessingen, lokalitet", "bebygget", "Bebyggelsesnavn.", ""),
    ("Waadt (Kanton)", "Kanton Vaud/Waadt", "admreg", "Kanton = administrativ enhed.", ""),
    ("Breitenburg", "Schloss/Gut Breitenburg", "anlaeg", "Herregård/fæstet renæssancesæde.", "[7]"),
    ("Chåtdet", "Sandsynlig fejl-/OCR-form, mulig Châtelet", "anlaeg", "USIKKER: intern registerform findes, men ekstern identifikation ikke sikkert verificeret.", ""),
    ("Genova", "Genova/Genoa", "bebygget", "By.", ""),
    ("Albano Laziale", "Italiensk by", "bebygget", "By.", ""),
    ("Ferrara", "Italiensk by", "bebygget", "By.", ""),
    ("Sachsen (Kongerige)", "Kongeriget Sachsen", "admreg", "Historisk stat/administrativ enhed.", ""),
]

assert len(VERIFIED_80) == 80, f"Expected 80 rows, got {len(VERIFIED_80)}"


def build_workbook(sv14_rows, out_path):
    wb = Workbook()

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    uncertain_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    title_font = Font(bold=True, size=13)
    wrap = Alignment(wrap_text=True, vertical="top")

    # --- Sheet 0: Læsevejledning ---
    ws0 = wb.active
    ws0.title = "Læsevejledning"
    ws0.column_dimensions["A"].width = 100
    lines = [
        ("Stedtypologi — klassifikationsarbejdsark", title_font),
        ("", None),
        ("Kilder:", Font(bold=True)),
        ("• Ark 2 (\"80 verificerede cases\"): den copilot-producerede klassifikation af 80-stedssamplet,"
         " arkiveret i docs/data-model/place-categorization-copilot-report.md.", None),
        ("• Ark 3 (\"SV14-register (481)\"): samtlige 481 poster i svNames/data/registers/places.xml,"
         " klassificeret automatisk efter beslutningsproceduren i docs/data-model/place-typology.md afsnit G.", None),
        ("• Ark 4 (\"Statistik\"): kategorifordeling for begge datasæt.", None),
        ("", None),
        ("Kategorier (slug -> navn):", Font(bold=True)),
        ("1  bebygget  — Bebyggede områder", None),
        ("2  admreg    — Lande, administrative enheder og regioner", None),
        ("3  vand      — Vandområder", None),
        ("4  landskab  — Landskabsformer og naturfænomener (inkl. øer)", None),
        ("5  anlaeg    — Bygninger, anlæg og fortidsminder", None),
        ("6  park      — Parker, haver og naturområder", None),
        ("   usikker   — kategori kan ikke afgøres (kun brugt for SV14-poster med uafklaret identitet)", None),
        ("", None),
        ("Selvkritik / metodenote (SV14-arket):", Font(bold=True)),
        ("Første automatiseringsforsøg lod alle 90 poster med den uspecifikke GeoNames-placeholder"
         " \"PPL\" (uden klassepræfiks) falde til default-kategorien \"bebygget\". Stikprøvekontrol af de"
         " faktiske 90 navne viste, at dette var forkert: PPL-koden bruges i dette register næsten"
         " udelukkende til kirker, slotte, teatre, monumenter og landskabsformer, som en"
         " GeoNames-matchning ikke er lykkedes for — ikke til bebyggelser. Efter individuel gennemgang"
         " af alle 90 navne (se PPL_NAME_OVERRIDE i byggescriptet) matcher SV14-fordelingen nu den"
         " oprindelige manuelle optælling i place-typology.md afsnit A næsten præcist"
         " (bebygget 209, admreg 15, vand 31, park 12 — alle eksakte match; landskab 73 og anlæg 139"
         " ligger tæt på de oprindelige ca.-skøn på hhv. 72 og 153).", None),
        ("", None),
        ("Kolonnen \"Metode\" i SV14-arket viser, hvordan hver klassifikation blev nået"
         " (GeoNames-klasse, kodeundtagelse, navngivet enkelttilfælde, eller individuel gennemgang af"
         " en PPL-post) — brug den til at prioritere stikprøvekontrol.", None),
        ("Gulmarkerede rækker er USIKRE og bør gennemgås manuelt før data skrives til det levende register.", None),
    ]
    for text, font in lines:
        ws0.append([text])
        cell = ws0.cell(row=ws0.max_row, column=1)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        if font:
            cell.font = font

    # --- Sheet 1: 80 verificerede cases ---
    ws1 = wb.create_sheet("80 verificerede cases")
    headers1 = ["Nr", "Stednavn", "Mest sandsynlige referent", "Kategori nr.",
                "Kategori (slug)", "Kategori (navn)", "Begrundelse", "Fodnote", "Usikker"]
    ws1.append(headers1)
    for col in range(1, len(headers1) + 1):
        c = ws1.cell(row=1, column=col)
        c.font = header_font
        c.fill = header_fill
        c.alignment = wrap

    for i, (name, referent, slug, reason, footnote) in enumerate(VERIFIED_80, start=1):
        is_uncertain = "JA" if reason.strip().upper().startswith("USIKKER") else ""
        row = [i, name, referent, CATEGORY_NUM.get(slug, ""), slug,
               CATEGORY_NAMES.get(slug, slug), reason, footnote, is_uncertain]
        ws1.append(row)
        if is_uncertain:
            for col in range(1, len(headers1) + 1):
                ws1.cell(row=ws1.max_row, column=col).fill = uncertain_fill

    widths1 = [5, 26, 42, 12, 14, 40, 55, 9, 9]
    for idx, w in enumerate(widths1, start=1):
        ws1.column_dimensions[get_column_letter(idx)].width = w
    ws1.freeze_panes = "A2"

    # --- Sheet 2: SV14-registret (fuld klassifikation) ---
    ws2 = wb.create_sheet("SV14-register (481)")
    headers2 = ["xml:id", "Stednavn", "Land", "Region", "GeoNames @type",
                "Kategori nr.", "Kategori (slug)", "Kategori (navn)", "Metode", "Begrundelse"]
    ws2.append(headers2)
    for col in range(1, len(headers2) + 1):
        c = ws2.cell(row=1, column=col)
        c.font = header_font
        c.fill = header_fill
        c.alignment = wrap

    for r in sv14_rows:
        row = [r['xml_id'], r['Stednavn'], r['Land'], r['Region'], r['GeoNames @type'],
               r['Kategori nr.'], r['Kategori (slug)'], r['Kategori (navn)'],
               r['Metode'], r['Begrundelse']]
        ws2.append(row)
        if r['Kategori (slug)'] == 'usikker':
            for col in range(1, len(headers2) + 1):
                ws2.cell(row=ws2.max_row, column=col).fill = uncertain_fill

    widths2 = [14, 26, 16, 20, 16, 12, 14, 38, 26, 55]
    for idx, w in enumerate(widths2, start=1):
        ws2.column_dimensions[get_column_letter(idx)].width = w
    ws2.freeze_panes = "A2"

    # --- Sheet 3: Statistik ---
    ws3 = wb.create_sheet("Statistik")
    ws3.append(["Datasæt", "Kategori", "Antal", "Andel"])
    for col in range(1, 5):
        c = ws3.cell(row=1, column=col)
        c.font = header_font
        c.fill = header_fill

    # 80-case stats. Kategori-optælling er uafhængig af usikkerhedsflaget:
    # Falkensten og Chåtdet tæller med under "Bygninger, anlæg og
    # fortidsminder" (deres tildelte kategori i kilderapporten), og
    # rapporteres separat i USIKKER-linjen som et overlejret flag, ikke
    # som en 7. pseudo-kategori.
    c80 = {}
    n_uncertain_80 = 0
    for _, _, slug, reason, _ in VERIFIED_80:
        c80[slug] = c80.get(slug, 0) + 1
        if reason.strip().upper().startswith("USIKKER"):
            n_uncertain_80 += 1
    order = ['bebygget', 'admreg', 'vand', 'landskab', 'anlaeg', 'park']
    for slug in order:
        n = c80.get(slug, 0)
        ws3.append(["80 verificerede cases", CATEGORY_NAMES[slug], n, f"{n/80*100:.1f}%"])
    ws3.append(["80 verificerede cases", "I ALT", 80, "100.0%"])
    ws3.append(["80 verificerede cases", "— heraf USIKKER (identifikation, ikke kategori)", n_uncertain_80, f"{n_uncertain_80/80*100:.1f}%"])

    ws3.append([])

    # SV14 stats. Her ER "usikker" en reel tilstand (identitet uafklaret,
    # jf. afsnit C: Falleberthor og den fejlkoblede Lilienstein-post) —
    # ikke kun et overlejret flag som for de 80 cases.
    from collections import Counter
    c_sv14 = Counter(r['Kategori (slug)'] for r in sv14_rows)
    total_sv14 = len(sv14_rows)
    order_sv14 = order + ['usikker']
    for slug in order_sv14:
        n = c_sv14.get(slug, 0)
        ws3.append(["SV14-register", CATEGORY_NAMES[slug], n, f"{n/total_sv14*100:.1f}%"])
    ws3.append(["SV14-register", "I ALT", total_sv14, "100.0%"])

    widths3 = [24, 40, 10, 10]
    for idx, w in enumerate(widths3, start=1):
        ws3.column_dimensions[get_column_letter(idx)].width = w

    wb.save(out_path)
    print(f"Saved: {out_path}")
    print(f"SV14 rows: {len(sv14_rows)}")
    print("SV14 category distribution:", dict(c_sv14))


if __name__ == '__main__':
    # svNames er et separat repo, normalt checket ud som søsterkatalog til
    # hca-open-repo. Overstyr med miljøvariablen SVNAMES_PLACES_XML, hvis
    # din lokale opsætning afviger.
    import os
    here = os.path.dirname(os.path.abspath(__file__))
    places_xml = os.environ.get(
        'SVNAMES_PLACES_XML',
        os.path.join(here, '..', '..', '..', 'svNames', 'data', 'registers', 'places.xml'),
    )
    out_xlsx = os.path.join(here, '..', '..', 'docs', 'data-model', 'exports',
                             'place-categorization-80cases-and-sv14register.xlsx')
    sv14_rows = load_sv14(places_xml)
    build_workbook(sv14_rows, out_xlsx)
