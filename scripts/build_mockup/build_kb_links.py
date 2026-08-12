#!/usr/bin/env python3
"""
build_kb_links.py
-----------------
Parses raw/1-KBDiaryLinkData-PQ-links-active.xlsm into
data/normalized/kb_diary_links.csv — one row per diary page with the
permanent link to Det Kgl. Biblioteks facsimile/EPUB edition.

The workbook has two sheets:

  KBLinks    Volumen · Page · KBLinkString   (4.413 rows, vols I–X)
  OffSetTab  VolumenRoman · VolumenString · PageOffset · NoOfPages

OffSetTab is the rule the links are generated from. KB's file names carry
a running sequence number that includes the volume's front matter, so a
diary page maps to

    hcadag{VV}_{offset + page - 1:03d}_{page}.xhtml

That reproduces 4.412 of the 4.413 rows exactly. The single exception is
vol I page 13, whose stored link points at page 32's file — the only
duplicated URL in the sheet, with both neighbours following the rule. It
is treated as a transcription slip in the source workbook: the computed
link is written instead and the discrepancy is reported below, so the
decision stays visible rather than buried. Fix the workbook and rerun to
make the warning disappear.

Vol XI is absent from the workbook (KB has published vols I–X), so its
56 pages get no link; the site simply omits the link on those pages.

Run from the repo root:
  python scripts/build_mockup/build_kb_links.py
"""

import csv
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl required:  pip install openpyxl")

REPO_ROOT = Path(__file__).resolve().parents[2]
SRC_XLSM  = REPO_ROOT / "raw" / "1-KBDiaryLinkData-PQ-links-active.xlsm"
OUT_CSV   = REPO_ROOT / "data" / "normalized" / "kb_diary_links.csv"

BASE = "https://epub3.kb.dk/hcadag/epub3/EPUB/"

VOL_NUM = {
    "I": 1, "II": 2, "III": 3, "IV": 4, "V": 5, "VI": 6,
    "VII": 7, "VIII": 8, "IX": 9, "X": 10, "XI": 11,
}


def pid(vol: str, page: int) -> str:
    """Same page id the rest of the mockup uses (build_diary_index.pid)."""
    return f"Pag{VOL_NUM.get(vol, 0):02d}{int(page):04d}"


def main() -> None:
    if not SRC_XLSM.exists():
        sys.exit(f"missing source workbook: {SRC_XLSM}")

    wb = openpyxl.load_workbook(SRC_XLSM, data_only=True)

    offsets = {}
    for r in wb["OffSetTab"].iter_rows(min_row=2, values_only=True):
        if r[0]:
            offsets[r[0]] = {"vv": str(r[1]), "offset": int(r[2]), "pages": int(r[3])}

    rows, mismatches = [], []
    for r in wb["KBLinks"].iter_rows(min_row=2, values_only=True):
        if not r[0]:
            continue
        vol, page = r[0], int(r[1])
        stored = (r[2] or "").strip()          # 8 cells carry a trailing space
        o = offsets.get(vol)
        if not o:
            sys.exit(f"volume {vol!r} in KBLinks has no OffSetTab entry")
        computed = f"{BASE}hcadag{o['vv']}_{o['offset'] + page - 1:03d}_{page}.xhtml"

        if stored == computed:
            url, source = stored, "workbook"
        else:
            url, source = computed, "computed"
            mismatches.append((vol, page, stored, computed))

        rows.append({
            "pag_id": pid(vol, page),
            "vol": vol,
            "page": page,
            "kb_url": url,
            "source": source,
        })

    rows.sort(key=lambda x: (VOL_NUM.get(x["vol"], 99), x["page"]))

    OUT_CSV.parent.mkdir(parents=True, exist_ok=True)
    with OUT_CSV.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=["pag_id", "vol", "page", "kb_url", "source"])
        w.writeheader()
        w.writerows(rows)

    per_vol = {}
    for r in rows:
        per_vol[r["vol"]] = per_vol.get(r["vol"], 0) + 1

    print(f"Read  {SRC_XLSM.relative_to(REPO_ROOT)}")
    print(f"Wrote {OUT_CSV.relative_to(REPO_ROOT)}  ({len(rows)} links)")
    print("  per volume: " + "  ".join(f"{v}={n}" for v, n in per_vol.items()))

    if mismatches:
        print(f"\n  ⚠  {len(mismatches)} row(s) disagree with the OffSetTab rule.")
        print("     The computed link was written; verify against the workbook:")
        for vol, page, stored, computed in mismatches:
            print(f"       vol {vol} p.{page}")
            print(f"         workbook: {stored.rsplit('/', 1)[-1]}")
            print(f"         computed: {computed.rsplit('/', 1)[-1]}")

    missing = sorted(set(offsets) - set(per_vol))
    if missing:
        print(f"\n  note: no links for vol(s) {', '.join(missing)}")
    print("Done.")


if __name__ == "__main__":
    main()
