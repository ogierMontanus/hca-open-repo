#!/usr/bin/env python3
"""
build_review_workbook.py
--------------------------
Builds a manual-review copy of data/parsed/personregister_xi_parsed.tsv
carrying the FULL content of every row, with the individual cells that
need a human decision highlighted.

Two outputs, because TSV cannot carry formatting:
  data/curated/personregister_xi_review_full.xlsx
      every column, frozen header, autofilter, colour-coded cells, and
      a per-row "review_flags" column naming the issues found.
  data/curated/personregister_xi_review_full.tsv
      the same data and flags as plain text, for grep/diff/pipeline use.

Flag classes (colour in the workbook):
  fused_entry      RED     -- 04_given_names still carries a whole second
                             entry: a reference run ("III 375.") or a
                             ", se: ..." cross-reference followed by the
                             next person's name. These are splits that
                             were reviewed and left, or that no cue caught.
  paren_unbalanced ORANGE  -- 13_raw_text has unmatched parentheses: the
                             tell-tale of a lost entry boundary, though
                             several are "enten X eller Y" constructions
                             the register itself never closes.
  hyphen_linewrap  YELLOW  -- an OCR line-wrap hyphen ("Gottholdi- ne").
                             The 884 mechanical cases are already fixed;
                             what remains here are the ones that must NOT
                             be joined ("Silke- og Klaedehandler",
                             "tysk- svensk"), listed so the decision is
                             visible rather than silently skipped.
  death_before_birth RED   -- 07 < 06 with no BC note; known source defects.
  no_refs_no_see   GREY    -- neither page references nor a see-target:
                             an entry with nothing to point at.
  suspect_years    ORANGE  -- a year parenthesis left inside the name
                             fields, so 04_given_names duplicates 06/07
                             ("E. A. (doed 1891)").

  python scripts/parsers/build_review_workbook.py
"""
import csv
import os
import re

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", ".."))
PARSED_TSV = os.path.join(ROOT, "data", "parsed", "personregister_xi_parsed.tsv")
OUT_XLSX = os.path.join(ROOT, "data", "curated", "personregister_xi_review_full.xlsx")
OUT_TSV = os.path.join(ROOT, "data", "curated", "personregister_xi_review_full.tsv")

FILL = {
    "red": PatternFill("solid", fgColor="FFC7CE"),
    "orange": PatternFill("solid", fgColor="FFD9A0"),
    "yellow": PatternFill("solid", fgColor="FFF2A8"),
    "grey": PatternFill("solid", fgColor="DDDDDD"),
}

REF_FUSED = re.compile(r"(?:I{1,3}|IV|VI{0,3}|IX|X)\s[\d\s\-]*\d\.\s+\S")
SEE_FUSED = re.compile(r",\s*se(?:\s+denne|:)?\s", re.IGNORECASE)
HYPHEN_WRAP = re.compile(r"[A-Za-zÀ-ÖØ-öø-ÿ]+-\s[a-zæøåàáâäéèêëíìîïóòôöúùûüçñ]+")
YEARS_IN_NAME = re.compile(r"\((?:ca\.\s*)?(?:d\.|død)?\s*\d{3,4}")
# Description-fusion marker a): a reference run followed by a new
# name-head, but INSIDE 09_description rather than 04_given_names --
# the class found via the second-pass review (Agier/Attila/Barfoed...).
DESC_FUSED_REF = re.compile(
    r"(?:I{1,3}|IV|VI{0,3}|IX|X)\s[\d\s\-]*\d\.\s+"
    r"(?=[A-ZÆØÅÖÜ][a-zæøåöäü]+(?:,|\s\(|\s[A-ZÆØÅÖÜ][a-zæøåöäü]+\s\())"
)
# A further person's own "Surname, Given (birth-death)," embedded
# anywhere deeper inside 09_description -- the class found by checking
# long cells and dash-heavy cells directly (Melchior/Hauch/Wulff/Collin).
EMBEDDED_NAME_YEAR = re.compile(
    r"[A-ZÆØÅÖÜ][a-zæøåöäü]+(?:\s\([^)]*\))?,\s"
    r"[A-ZÆØÅÖÜ][\wæøåöäü.]*(?:\s[A-ZÆØÅÖÜ]?[\wæøåöäü.]*)*\s?"
    r"\((?:ca\.\s*)?(?:d\.|død)?\s*\d{3,4}[\s–—\-]*\d{0,4}\),"
)


def paren_balance(s):
    return s.count("(") - s.count(")")


def flag_row(r):
    """Return {column name: [(flag, colour), ...]} for one row."""
    flags = {}

    def add(col, name, colour):
        flags.setdefault(col, []).append((name, colour))

    gn = r["04_given_names"]
    if REF_FUSED.search(gn):
        add("04_given_names", "fused_entry:reference_run", "red")
    if SEE_FUSED.search(gn):
        add("04_given_names", "fused_entry:see_reference", "red")
    if YEARS_IN_NAME.search(gn):
        add("04_given_names", "suspect_years:year_left_in_name", "orange")

    if paren_balance(r["13_raw_text"]):
        add("13_raw_text", "paren_unbalanced", "orange")

    if DESC_FUSED_REF.search(r["09_description"]):
        add("09_description", "fused_entry:description_reference_run", "red")

    # Skip position 0 -- that's this row's own name, already parsed.
    if any(m.start() > 20 for m in EMBEDDED_NAME_YEAR.finditer(r["09_description"])):
        add("09_description", "fused_entry:embedded_name_year", "red")

    for col in ("03_surname", "04_given_names", "09_description"):
        if HYPHEN_WRAP.search(r[col]):
            add(col, "hyphen_linewrap:verify_do_not_join", "yellow")

    b, d = r["06_birth_year"], r["07_death_year"]
    if b and d and int(d) < int(b) and "f. Kr" not in r["08_year_note"]:
        add("06_birth_year", "death_before_birth", "red")
        add("07_death_year", "death_before_birth", "red")

    if not r["11_references_parsed"] and not r["12_see_also"]:
        add("11_references_parsed", "no_refs_no_see", "grey")

    return flags


def main():
    with open(PARSED_TSV, encoding="utf-8") as f:
        rows = list(csv.DictReader(f, delimiter="\t"))
    cols = list(rows[0].keys())
    out_cols = cols + ["review_flags"]

    wb = Workbook()
    ws = wb.active
    ws.title = "personregister XI"

    header_font = Font(bold=True)
    for j, name in enumerate(out_cols, start=1):
        c = ws.cell(row=1, column=j, value=name)
        c.font = header_font
        c.alignment = Alignment(vertical="top")

    n_flagged_rows = 0
    flag_census = {}
    tsv_rows = []

    for i, r in enumerate(rows, start=2):
        flags = flag_row(r)
        all_names = sorted({n for lst in flags.values() for n, _ in lst})
        for n in all_names:
            flag_census[n] = flag_census.get(n, 0) + 1
        if all_names:
            n_flagged_rows += 1

        for j, name in enumerate(cols, start=1):
            cell = ws.cell(row=i, column=j, value=r[name])
            cell.alignment = Alignment(vertical="top", wrap_text=False)
            if name in flags:
                names = [n for n, _ in flags[name]]
                # Most severe colour wins: red > orange > yellow > grey.
                order = ["red", "orange", "yellow", "grey"]
                colour = min((c for _, c in flags[name]), key=order.index)
                cell.fill = FILL[colour]
                cell.comment = Comment("; ".join(names), "review")

        ws.cell(row=i, column=len(out_cols), value="; ".join(all_names))

        tsv_rows.append({**r, "review_flags": "; ".join(all_names)})

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(out_cols))}{len(rows) + 1}"
    widths = {
        "01_entry_id": 12, "02_entry_type": 16, "03_surname": 24,
        "04_given_names": 46, "05_sort_key": 40, "06_birth_year": 7,
        "07_death_year": 7, "08_year_note": 14, "09_description": 60,
        "10_references_raw": 34, "11_references_parsed": 34,
        "12_see_also": 26, "13_raw_text": 90, "review_flags": 40,
    }
    for j, name in enumerate(out_cols, start=1):
        ws.column_dimensions[get_column_letter(j)].width = widths.get(name, 18)

    # Legend on a second sheet, so the colours are self-explanatory.
    legend = wb.create_sheet("legend")
    legend.append(["colour", "flag", "meaning", "rows"])
    legend["A1"].font = legend["B1"].font = header_font
    legend["C1"].font = legend["D1"].font = header_font
    meanings = [
        ("red", "fused_entry:reference_run",
         "04_given_names still contains a following entry after a page-reference run", ),
        ("red", "fused_entry:see_reference",
         "04_given_names still contains a following entry after a ', se:' cross-reference"),
        ("red", "fused_entry:description_reference_run",
         "09_description still contains a following entry after ITS OWN reference run "
         "(the row's own year-parenthesis parsed fine, but the description trails a "
         "second person, e.g. Agier/Attila/Barfoed) -- xlsx could not confirm a safe "
         "split for these, unlike the 47 already applied"),
        ("red", "fused_entry:embedded_name_year",
         "a further person's own 'Surname, Given (years),' is buried deeper inside "
         "09_description, often after a ', se: X' cross-reference (Hauch/Wulff/Collin-"
         "shaped) -- xlsx could not confirm a safe split for these, unlike the 132 "
         "already applied"),
        ("red", "death_before_birth",
         "death year precedes birth year (no BC note) -- known source/OCR defects"),
        ("orange", "paren_unbalanced",
         "13_raw_text has unmatched parentheses -- often a lost entry boundary, "
         "but 'enten X eller Y' entries are unbalanced in the source too"),
        ("orange", "suspect_years:year_left_in_name",
         "a year parenthesis is still inside 04_given_names, duplicating 06/07"),
        ("yellow", "hyphen_linewrap:verify_do_not_join",
         "hyphen + space: the 884 real line-wraps are already joined; these are the "
         "ones that must stay split ('Silke- og ...', 'tysk- svensk')"),
        ("grey", "no_refs_no_see",
         "entry has neither page references nor a see-target"),
    ]
    for colour, flag, meaning in meanings:
        legend.append(["", flag, meaning, flag_census.get(flag, 0)])
        legend.cell(row=legend.max_row, column=1).fill = FILL[colour]
    for col, w in (("A", 10), ("B", 38), ("C", 96), ("D", 8)):
        legend.column_dimensions[col].width = w

    os.makedirs(os.path.dirname(OUT_XLSX), exist_ok=True)

    # The TSV is master1 — the machine-readable source the rest of the
    # pipeline reads. Write it FIRST: the .xlsx is only the review surface,
    # and having it open in Excel locks the file on Windows. Writing the
    # workbook first meant a spreadsheet left open silently blocked the
    # master from being rebuilt at all.
    with open(OUT_TSV, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=out_cols, delimiter="\t")
        w.writeheader()
        w.writerows(tsv_rows)

    print(f"rows: {len(rows)}   flagged rows: {n_flagged_rows}")
    for k in sorted(flag_census):
        print(f"  {k:42s} {flag_census[k]}")
    print(f"wrote {os.path.relpath(OUT_TSV, ROOT)}")

    try:
        wb.save(OUT_XLSX)
        print(f"wrote {os.path.relpath(OUT_XLSX, ROOT)}")
    except PermissionError:
        print(f"[!] kunne IKKE skrive {os.path.relpath(OUT_XLSX, ROOT)} "
              "— filen er åben i Excel. TSV'en er opdateret; luk regnearket "
              "og kør igen for at opdatere gennemsynsarket.")
        return 1
    return 0


if __name__ == "__main__":
    main()
