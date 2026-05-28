#!/usr/bin/env python3
"""Convert an Excel (.xlsx) file to a UTF-8 tab-separated values (.tsv) file.

Reads computed cell values (data_only=True), not formulas. The active
worksheet is used; pass --sheet to pick a specific one. Empty cells are
written as empty strings. Encoding: UTF-8, no BOM.

Usage:
    python xlsx_to_tsv.py <input.xlsx> [output.tsv] [--sheet NAME]
"""

import argparse
import csv
import pathlib

import openpyxl


def xlsx_to_tsv(
    src: pathlib.Path,
    dst: pathlib.Path | None = None,
    sheet: str | None = None,
) -> pathlib.Path:
    dst_path = dst or src.with_suffix(".tsv")

    wb = openpyxl.load_workbook(src, data_only=True)
    ws = wb[sheet] if sheet else wb.active

    with open(dst_path, "w", encoding="utf-8", newline="") as f:
        writer = csv.writer(f, delimiter="\t")
        for row in ws.iter_rows(values_only=True):
            writer.writerow(["" if v is None else str(v) for v in row])

    print(f"Converted: {src.name}  [sheet={ws.title!r}]")
    print(f"      To:  {dst_path}")
    print(f"  Columns: {ws.max_column}  |  Rows: {ws.max_row}")
    return dst_path


def main() -> None:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("input", type=pathlib.Path)
    ap.add_argument("output", type=pathlib.Path, nargs="?", default=None)
    ap.add_argument("--sheet", default=None, help="Sheet name (default: active)")
    args = ap.parse_args()
    xlsx_to_tsv(args.input, args.output, args.sheet)


if __name__ == "__main__":
    main()
